import openpyxl


class Excel(object):

    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheet_name = sheet_name

    def excel_read(self):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheet_name]
        sheet_data = list(sheet.rows)
        # title = []
        # for i in sheet_data[0]:
        #     title.append(i.value)
        title = [i.value for i in sheet_data[0]]
        # print(title)
        case_data = []
        for i in sheet_data[1:]:
            data = [j.value for j in i]
            case_data.append(dict(zip(title, data)))
            # print(data)
        return case_data

    # 写入excel文件
    def write_data(self, row, column, value):
        workbook = openpyxl.load_workbook(filename=self.filename)
        sheet = workbook[self.sheet_name]
        sheet.cell(row=row, column=column, value=value)
        workbook.save(self.filename)

